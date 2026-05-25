import sqlite3
import os
import openpyxl
from database import get_db_connection
from excel_export import export_to_excel
from config import BASE_DIR

def run_tests():
    print("Starting automated tests...")
    
    # 1. Insert dummy records
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Clear old test data if any
    cursor.execute("DELETE FROM consumers WHERE name LIKE '%TEST_USER%'")
    conn.commit()
    
    # Insert test record 1 (Unmetered)
    dummy_1 = {
        'name': 'TEST_USER_ONE',
        'customer_group': '01 - DOMESTIC',
        'tel1': '10',
        'tel2': '20',
        'fax': '30',
        'mobile_phone': '9999999991',
        'email': 'test1@bpcl.com',
        'application_form_no': '10-4609-36-TEST-001',
        'locality': 'Danipali',
        'aadhaar1': '1111',
        'aadhaar2': '2222',
        'aadhaar3': '3333',
        'aadhar_name': 'TEST USER ONE',
        'contact_title': 'Mr',
        'first_name': 'TEST_USER',
        'middle_name': '',
        'last_name': 'ONE',
        'father_name': 'FATHER ONE',
        'address': 'H.No: 101, Danipali, Sambalpur, 768004',
        'date_of_birth': '01.01.1990',
        'gender': 'Male',
        'profession': 'Service',
        'house_no': '101',
        'street_no': 'Main Road',
        'postal_code': '768004',
        'city': 'Sambalpur',
        'type_of_property': 'ROW HOUSE',
        'type_of_ownership': 'SELF OWNED',
        'pan': 'ABCDE1234F',
        'scheme_code': 'PNG-DOM-08',
        'cheque_no': '123456',
        'cheque_date': '20.05.2026',
        'cheque_received_date': '21.05.2026',
        'cheque_amount': '500.0',
        'bank_name': 'State Bank of India',
        'meter_no': '' # Empty/unmetered
    }
    
    # Insert test record 2 (Metered)
    dummy_2 = {
        'name': 'TEST_USER_TWO',
        'customer_group': '01 - DOMESTIC',
        'tel1': '11',
        'tel2': '22',
        'fax': '33',
        'mobile_phone': '9999999992',
        'email': 'test2@bpcl.com',
        'application_form_no': '10-4609-36-TEST-002',
        'locality': 'Bhalupali',
        'aadhaar1': '5555',
        'aadhaar2': '6666',
        'aadhaar3': '7777',
        'aadhar_name': 'TEST USER TWO',
        'contact_title': 'Mrs',
        'first_name': 'TEST_USER',
        'middle_name': '',
        'last_name': 'TWO',
        'father_name': 'FATHER TWO',
        'address': 'H.No: 202, Bhalupali, Sambalpur, 768004',
        'date_of_birth': '02.02.1985',
        'gender': 'Female',
        'profession': 'Business',
        'house_no': '202',
        'street_no': 'Sub Road',
        'postal_code': '768004',
        'city': 'Sambalpur',
        'type_of_property': 'APARTMENT',
        'type_of_ownership': 'RENTED',
        'pan': 'XYZWP9876Q',
        'scheme_code': 'PNG-DOM-08',
        'cheque_no': '654321',
        'cheque_date': '19.05.2026',
        'cheque_received_date': '20.05.2026',
        'cheque_amount': '1000.0',
        'bank_name': 'HDFC Bank',
        'meter_no': 'RR2301099999' # Metered
    }
    
    def insert_record(d):
        keys = list(d.keys())
        cols = ', '.join(keys)
        placeholders = ', '.join('?' for _ in keys)
        cursor.execute(f"INSERT INTO consumers ({cols}) VALUES ({placeholders})", [d[k] for k in keys])
        return cursor.lastrowid
        
    id1 = insert_record(dummy_1)
    id2 = insert_record(dummy_2)
    conn.commit()
    print(f"Inserted test records: ID {id1} and ID {id2}")
    
    # 2. Run Export
    export_file = os.path.join(BASE_DIR, 'data', 'test_verification_export.xlsx')
    if os.path.exists(export_file):
        os.remove(export_file)
        
    print(f"Running Excel export to: {export_file}")
    count = export_to_excel(export_file, [id1, id2])
    print(f"Export completed: {count} records written.")
    
    # 3. Verify Excel File Columns & Values
    print("Verifying excel file contents...")
    wb = openpyxl.load_workbook(export_file)
    ws = wb['Sheet1']
    
    # Verify Header
    header_col_50 = ws.cell(row=1, column=50).value
    print(f"Header in Col 50 (AX): {header_col_50} (Expected: Meter No.)")
    assert header_col_50 == "Meter No.", "Header verification failed!"
    
    # Verify Row 2 values (TEST_USER_ONE)
    name_row2 = ws.cell(row=2, column=1).value
    form_row2 = ws.cell(row=2, column=8).value
    meter_row2 = ws.cell(row=2, column=50).value
    print(f"Row 2: Name='{name_row2}', Form='{form_row2}', Meter='{meter_row2}'")
    assert name_row2 == "TEST_USER_ONE", f"Name mismatch: {name_row2}"
    assert form_row2 == "10-4609-36-TEST-001", f"Form mismatch: {form_row2}"
    assert meter_row2 == "" or meter_row2 is None, f"Meter mismatch: {meter_row2}"
    
    # Verify Row 3 values (TEST_USER_TWO)
    name_row3 = ws.cell(row=3, column=1).value
    form_row3 = ws.cell(row=3, column=8).value
    meter_row3 = ws.cell(row=3, column=50).value
    print(f"Row 3: Name='{name_row3}', Form='{form_row3}', Meter='{meter_row3}'")
    assert name_row3 == "TEST_USER_TWO", f"Name mismatch: {name_row3}"
    assert form_row3 == "10-4609-36-TEST-002", f"Form mismatch: {form_row3}"
    assert meter_row3 == "RR2301099999", f"Meter mismatch: {meter_row3}"
    
    # Clean up test database records
    cursor.execute("DELETE FROM consumers WHERE name LIKE '%TEST_USER%'")
    conn.commit()
    conn.close()
    
    wb.close()
    print("ALL TESTS PASSED SUCCESSFULLY!")

if __name__ == '__main__':
    run_tests()
