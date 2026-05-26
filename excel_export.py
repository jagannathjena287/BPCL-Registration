import openpyxl
import os
import shutil
import sqlite3
from database import get_db_connection, get_db_cursor, execute_query
from config import BASE_DIR

TEMPLATE_FILE = os.path.join(BASE_DIR, 'Copy of BP Utility Template 64 Bit.xlsx')

def export_to_excel(output_filepath, consumer_ids=None):
    """
    Copies the BPCL utility template, loads it, writes consumer registrations from database to Sheet1,
    adds the 'Meter No.' column header at Column AX, and saves it.
    """
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"Template file not found at {TEMPLATE_FILE}")
        
    # Copy the template to the output file path
    shutil.copyfile(TEMPLATE_FILE, output_filepath)
    
    # Load workbook and Sheet1
    wb = openpyxl.load_workbook(output_filepath)
    if 'Sheet1' not in wb.sheetnames:
        raise ValueError("Sheet1 not found in the Excel template.")
        
    ws = wb['Sheet1']
    
    # Let's ensure Column AX (index 50) has 'Meter No.' header
    ws.cell(row=1, column=50, value="Meter No.")
    
    # Fetch data from database
    conn = get_db_connection()
    cursor = get_db_cursor(conn)
    
    if consumer_ids:
        # Export only selected consumer IDs
        placeholders = ','.join('?' for _ in consumer_ids)
        query = f"SELECT * FROM consumers WHERE id IN ({placeholders}) ORDER BY id ASC"
        execute_query(cursor, query, consumer_ids)
    else:
        # Export all consumers
        execute_query(cursor, "SELECT * FROM consumers ORDER BY id ASC")
        
    rows = cursor.fetchall()
    conn.close()
    
    # Mapping of DB column names to spreadsheet column indices (1-based)
    mapping = {
        'name': 1,               # A: Name
        'customer_group': 2,      # B: Group
        'tel1': 3,               # C: Tel1
        'tel2': 4,               # D: Tel2
        'fax': 5,                # E: Fax
        'mobile_phone': 6,       # F: MobilePhone
        'email': 7,              # G: EMail
        'application_form_no': 8, # H: ApplicationFormNo
        'geographical_area': 9,  # I: GeoGraphicalArea
        'charge_area': 10,        # J: ChargeArea
        'locality': 11,          # K: Locality
        'aadhaar1': 12,          # L: Aadhaar1
        'aadhaar2': 13,          # M: Aadhaar2
        'aadhaar3': 14,          # N: Aadhaar3
        'aadhar_name': 15,        # O: AadharName
        'contact_title': 16,     # P: ContactTitle
        'first_name': 17,        # Q: FirstName
        'middle_name': 18,       # R: MiddleName
        'last_name': 19,         # S: LastName
        'father_name': 20,       # T: FatherName
        'designation': 21,       # U: Designation
        'address': 22,           # V: Address
        'telephone1': 23,        # W: Telephone1
        'telephone2': 24,        # X: Telephone2
        'contact_mobile': 25,    # Y: ContactMobile
        'contact_email': 26,     # Z: ContactEmail
        'date_of_birth': 27,     # AA: DateofBirth
        'gender': 28,            # AB: Gender
        'profession': 29,        # AC: Profession
        'house_no': 30,          # AD: HouseNo
        'street_no': 31,          # AE: StreetNo
        'po_box': 32,            # AF: POBox
        'postal_code': 33,       # AG: PostalCode
        'city': 34,              # AH: City
        'gstin': 35,             # AI: GSTIN
        'gst_type': 36,          # AJ: GSTType
        'type_of_property': 37,  # AK: TypeofProperty
        'type_of_ownership': 38, # AL: TypeofOwnership
        'pan': 39,               # AM: Pan
        'customer_no': 40,       # AN: CustomerNo
        'distributor_name': 41,  # AO: DistributorName
        'omc_name': 42,          # AP: OMCName
        'scheme_code': 43,       # AQ: SchemeCode
        'cheque_no': 44,         # AR: ChequeNo
        'cheque_date': 45,       # AS: ChequeDate
        'cheque_received_date': 46,# AT: ChequeReceivedDate
        'cheque_amount': 47,     # AU: ChequeAmount
        'bank_name': 48,         # AV: BankName
        # 49 (AW) is empty
        'meter_no': 50           # AX: Meter No.
    }
    
    # Start writing at Row 2
    current_row = 2
    for r in rows:
        row_dict = dict(r)
        for field, col_idx in mapping.items():
            val = row_dict.get(field)
            
            # Formatter or type cast if necessary (e.g. cheque amount to int/float if digits)
            if field == 'cheque_amount' and val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            
            ws.cell(row=current_row, column=col_idx, value=val)
        current_row += 1
        
    wb.save(output_filepath)
    wb.close()
    print(f"Exported {len(rows)} records to {output_filepath}")
    return len(rows)

if __name__ == '__main__':
    # Test export
    export_to_excel(os.path.join(BASE_DIR, 'data', 'test_export.xlsx'))
